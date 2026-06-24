from __future__ import annotations

import csv
from pathlib import Path

from engine.domain.project import MethodConfig, PerturbationConfig, ProjectConfig, WellConfig
from engine.io.project_loader import load_project_json as io_load_project_json
from engine.io.project_writer import write_project_json


def create_empty_project(name: str = "CERITANYA INI SIMULATOR") -> ProjectConfig:
	project = ProjectConfig(name=name)
	project.description = "Desktop reservoir simulation workspace."
	project.reference_data.reference_depth = 6500.0
	project.reference_data.reference_pressure = 2500.0
	project.reference_data.water_density_reference = 1.0
	project.grid_spec.nx = 5
	project.grid_spec.ny = 5
	project.grid_spec.nz = 1
	project.grid_spec.dx = 100.0
	project.grid_spec.dy = 100.0
	project.grid_spec.dz = 50.0
	project.initial_conditions.reference_depth = 6500.0
	project.initial_conditions.initial_sw = 0.25
	project.initial_conditions.initial_sg = 0.05
	project.solver.initial_timestep_days = 1.0
	project.solver.min_timestep_days = 0.05
	project.solver.max_time_days = 10.0
	project.solver.timestep_growth_factor = 1.1
	project.solver.timestep_shrink_factor = 0.5
	project.solver.max_step_retries = 8
	project.solver.max_newton_iterations = 50
	project.solver.residual_tolerance = 1e-8
	project.solver.parameter_tolerance_pressure = 1e-3
	project.solver.parameter_tolerance_saturation = 1e-4
	project.solver.residual_norm_floor = 0.1
	project.solver.newton_pressure_damping = 0.7
	project.solver.newton_saturation_damping = 0.7
	project.solver.max_pressure_correction = 10.0
	project.solver.max_saturation_correction = 0.001
	project.run.case_name = f"{name} Base Case"
	load_example_pvt_tables(project)
	load_example_rock_tables(project)
	mark_project_clean(project)
	return project


def update_project_metadata(
	project_config: ProjectConfig,
	*,
	name: str | None = None,
	description: str | None = None,
	case_name: str | None = None,
) -> ProjectConfig:
	if name is not None:
		project_config.name = name
	if description is not None:
		project_config.description = description
	if case_name is not None:
		project_config.run.case_name = case_name
	project_config.is_dirty = True
	return project_config


def update_grid_spec(
	project_config: ProjectConfig,
	*,
	nx: int,
	ny: int,
	nz: int,
	dx: float,
	dy: float,
	dz: float,
) -> ProjectConfig:
	project_config.grid_spec.nx = nx
	project_config.grid_spec.ny = ny
	project_config.grid_spec.nz = nz
	project_config.grid_spec.dx = dx
	project_config.grid_spec.dy = dy
	project_config.grid_spec.dz = dz
	project_config.is_dirty = True
	return project_config


def update_grid_connectivity(project_config: ProjectConfig, connectivity: int) -> ProjectConfig:
	project_config.grid_spec.connectivity = connectivity
	project_config.constraints.grid_confirmed = True
	project_config.is_dirty = True
	return project_config


def update_initial_conditions(
	project_config: ProjectConfig,
	*,
	reference_depth: float,
	initial_sw: float,
	initial_sg: float,
	reference_pressure: float,
) -> ProjectConfig:
	project_config.initial_conditions.reference_depth = reference_depth
	project_config.initial_conditions.initial_sw = initial_sw
	project_config.initial_conditions.initial_sg = initial_sg
	project_config.reference_data.reference_depth = reference_depth
	project_config.reference_data.reference_pressure = reference_pressure
	project_config.is_dirty = True
	return project_config


def update_solver_config(
	project_config: ProjectConfig,
	*,
	initial_timestep_days: float | None = None,
	min_timestep_days: float | None = None,
	max_time_days: float | None = None,
	timestep_growth_factor: float | None = None,
	timestep_shrink_factor: float | None = None,
	max_step_retries: int | None = None,
	max_newton_iterations: int | None = None,
	residual_tolerance: float | None = None,
	parameter_tolerance_pressure: float | None = None,
	parameter_tolerance_saturation: float | None = None,
	residual_norm_floor: float | None = None,
	newton_pressure_damping: float | None = None,
	newton_saturation_damping: float | None = None,
	max_pressure_correction: float | None = None,
	max_saturation_correction: float | None = None,
) -> ProjectConfig:
	if initial_timestep_days is not None:
		project_config.solver.initial_timestep_days = initial_timestep_days
	if min_timestep_days is not None:
		project_config.solver.min_timestep_days = min_timestep_days
	if max_time_days is not None:
		project_config.solver.max_time_days = max_time_days
	if timestep_growth_factor is not None:
		project_config.solver.timestep_growth_factor = timestep_growth_factor
	if timestep_shrink_factor is not None:
		project_config.solver.timestep_shrink_factor = timestep_shrink_factor
	if max_step_retries is not None:
		project_config.solver.max_step_retries = max_step_retries
	if max_newton_iterations is not None:
		project_config.solver.max_newton_iterations = max_newton_iterations
	if residual_tolerance is not None:
		project_config.solver.residual_tolerance = residual_tolerance
	if parameter_tolerance_pressure is not None:
		project_config.solver.parameter_tolerance_pressure = parameter_tolerance_pressure
	if parameter_tolerance_saturation is not None:
		project_config.solver.parameter_tolerance_saturation = parameter_tolerance_saturation
	if residual_norm_floor is not None:
		project_config.solver.residual_norm_floor = residual_norm_floor
	if newton_pressure_damping is not None:
		project_config.solver.newton_pressure_damping = newton_pressure_damping
	if newton_saturation_damping is not None:
		project_config.solver.newton_saturation_damping = newton_saturation_damping
	if max_pressure_correction is not None:
		project_config.solver.max_pressure_correction = max_pressure_correction
	if max_saturation_correction is not None:
		project_config.solver.max_saturation_correction = max_saturation_correction
	project_config.is_dirty = True
	return project_config


def load_example_pvt_tables(project_config: ProjectConfig) -> ProjectConfig:
	project_config.pvt_tables = {
		"bo": [(1000.0, 1.22), (2500.0, 1.08), (4000.0, 0.98)],
		"bw": [(1000.0, 1.03), (2500.0, 1.01), (4000.0, 0.99)],
		"bg": [(1000.0, 0.005), (2500.0, 0.003), (4000.0, 0.002)],
		"mu_o": [(1000.0, 2.1), (2500.0, 1.8), (4000.0, 1.5)],
		"mu_w": [(1000.0, 0.55), (2500.0, 0.58), (4000.0, 0.62)],
		"mu_g": [(1000.0, 0.02), (2500.0, 0.025), (4000.0, 0.03)],
		"rso": [(1000.0, 180.0), (2500.0, 320.0), (4000.0, 420.0)],
		"rsw": [(1000.0, 2.0), (2500.0, 2.5), (4000.0, 3.0)],
	}
	project_config.is_dirty = True
	return project_config


def import_pvt_tables_from_file(project_config: ProjectConfig, file_path: str | Path) -> ProjectConfig:
	path = Path(file_path)
	if not path.exists():
		raise ValueError(f"File tidak ditemukan: {path}")

	suffix = path.suffix.lower()
	if suffix == ".csv":
		rows = _read_rows_from_csv(path)
	elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
		rows = _read_rows_from_excel(path)
	else:
		raise ValueError("Format file tidak didukung. Gunakan .csv, .xlsx, atau .xls")

	tables = _rows_to_pvt_tables(rows)
	if not tables:
		raise ValueError("Tidak ada data PVT valid yang bisa di-import.")

	project_config.pvt_tables = tables
	project_config.is_dirty = True
	return project_config


def import_rock_tables_from_file(project_config: ProjectConfig, file_path: str | Path) -> ProjectConfig:
	path = Path(file_path)
	if not path.exists():
		raise ValueError(f"File tidak ditemukan: {path}")

	suffix = path.suffix.lower()
	if suffix == ".csv":
		rows = _read_rows_from_csv(path)
	elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
		rows = _read_rows_from_excel(path)
	else:
		raise ValueError("Format file tidak didukung. Gunakan .csv, .xlsx, atau .xls")

	tables = _rows_to_rock_tables(rows)
	if not tables:
		raise ValueError("Tidak ada data rock-fluid valid yang bisa di-import.")

	project_config.rock_tables = tables
	project_config.is_dirty = True
	return project_config


def clear_pvt_tables(project_config: ProjectConfig) -> ProjectConfig:
	project_config.pvt_tables.clear()
	project_config.is_dirty = True
	return project_config


def _read_rows_from_csv(path: Path) -> list[dict[str, object]]:
	with path.open("r", encoding="utf-8-sig", newline="") as handle:
		reader = csv.DictReader(handle)
		if not reader.fieldnames:
			raise ValueError("Header CSV kosong.")
		rows: list[dict[str, object]] = []
		for raw_row in reader:
			row: dict[str, object] = {}
			for key, value in raw_row.items():
				if key is None:
					continue
				row[_norm_col_name(key)] = "" if value is None else str(value).strip()
			if any(str(v).strip() for v in row.values()):
				rows.append(row)
		return rows


def _read_rows_from_excel(path: Path) -> list[dict[str, object]]:
	if path.suffix.lower() != ".xls":
		try:
			from openpyxl import load_workbook
		except ImportError:
			load_workbook = None
		if load_workbook is not None:
			wb = load_workbook(path, read_only=True, data_only=True)
			ws = wb.active
			rows_iter = ws.iter_rows(values_only=True)
			headers_raw = next(rows_iter, None)
			if headers_raw is None:
				raise ValueError("Sheet Excel kosong.")
			headers = [_norm_col_name(h) for h in headers_raw]
			rows: list[dict[str, object]] = []
			for excel_row in rows_iter:
				row: dict[str, object] = {}
				for idx, cell_value in enumerate(excel_row):
					if idx >= len(headers):
						continue
					col_name = headers[idx]
					if not col_name:
						continue
					row[col_name] = "" if cell_value is None else cell_value
				if any(str(v).strip() for v in row.values()):
					rows.append(row)
			return rows

	# Fallback for .xls or when openpyxl is not available
	try:
		import pandas as pd
	except ImportError as exc:
		raise ValueError(
			"Import Excel butuh package tambahan. Install openpyxl (untuk .xlsx) atau pandas/xlrd (untuk .xls)."
		) from exc

	df = pd.read_excel(path)
	rows = []
	for row_dict in df.to_dict(orient="records"):
		row: dict[str, object] = {}
		for key, value in row_dict.items():
			if key is None:
				continue
			norm_key = _norm_col_name(str(key))
			if value is None:
				row[norm_key] = ""
			elif isinstance(value, float) and str(value) == "nan":
				row[norm_key] = ""
			else:
				row[norm_key] = value
		if any(str(v).strip() for v in row.values()):
			rows.append(row)
	return rows


def _rows_to_pvt_tables(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	if not rows:
		return {}

	cols = set(rows[0].keys())
	long_mode = ({"table", "pressure", "value"} <= cols)
	wide_keys = {"bo", "bw", "bg", "mu_o", "mu_w", "mu_g", "rso", "rsw"}
	wide_mode = ("pressure" in cols and bool(wide_keys & cols))

	if long_mode:
		return _parse_long_rows(rows)
	if wide_mode:
		return _parse_wide_rows(rows)

	raise ValueError(
		"Format kolom tidak dikenali. Gunakan format long: table,pressure,value "
		"atau format wide: pressure,bo,bw,bg,mu_o,mu_w,mu_g,rso,rsw"
	)


def _rows_to_rock_tables(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	if not rows:
		return {}

	cols = set(rows[0].keys())
	long_mode = ({"table", "saturation", "value"} <= cols)
	wide_keys = {"kro", "krw", "pcow", "krg", "pcgw"}
	wide_mode = bool(wide_keys & cols) and bool({"saturation", "sw", "sg"} & cols)

	if long_mode:
		return _parse_long_rock_rows(rows)
	if wide_mode:
		return _parse_wide_rock_rows(rows)

	raise ValueError(
		"Format kolom rock-fluid tidak dikenali. Gunakan format long: table,saturation,value "
		"atau format wide dengan kolom saturation/sw/sg dan salah satu dari kro,krw,pcow,krg,pcgw"
	)


def _parse_long_rows(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	tables: dict[str, list[tuple[float, float]]] = {}
	for i, row in enumerate(rows, start=2):
		table_name = str(row.get("table", "")).strip().lower()
		if not table_name:
			raise ValueError(f"Kolom table kosong pada baris {i}.")
		pressure = _to_float(row.get("pressure"), "pressure", i)
		value = _to_float(row.get("value"), "value", i)
		tables.setdefault(table_name, []).append((pressure, value))

	for pairs in tables.values():
		pairs.sort(key=lambda item: item[0])
	return tables


def _parse_long_rock_rows(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	tables: dict[str, list[tuple[float, float]]] = {}
	for i, row in enumerate(rows, start=2):
		table_name = str(row.get("table", "")).strip().lower()
		if not table_name:
			raise ValueError(f"Kolom table kosong pada baris {i}.")
		saturation = _to_float(row.get("saturation"), "saturation", i)
		value = _to_float(row.get("value"), "value", i)
		tables.setdefault(table_name, []).append((saturation, value))

	for pairs in tables.values():
		pairs.sort(key=lambda item: item[0])
	return tables


def _parse_wide_rows(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	tables: dict[str, list[tuple[float, float]]] = {}
	ordered_tables = ["bo", "bw", "bg", "mu_o", "mu_w", "mu_g", "rso", "rsw"]

	for i, row in enumerate(rows, start=2):
		pressure = _to_float(row.get("pressure"), "pressure", i)
		for table_name in ordered_tables:
			raw_val = row.get(table_name)
			if raw_val is None or str(raw_val).strip() == "":
				continue
			value = _to_float(raw_val, table_name, i)
			tables.setdefault(table_name, []).append((pressure, value))

	for pairs in tables.values():
		pairs.sort(key=lambda item: item[0])
	return tables


def _parse_wide_rock_rows(rows: list[dict[str, object]]) -> dict[str, list[tuple[float, float]]]:
	tables: dict[str, list[tuple[float, float]]] = {}
	ordered_tables = ["kro", "krw", "pcow", "krg", "pcgw"]

	for i, row in enumerate(rows, start=2):
		raw_saturation = row.get("saturation")
		if raw_saturation is None or str(raw_saturation).strip() == "":
			raw_saturation = row.get("sw")
		if raw_saturation is None or str(raw_saturation).strip() == "":
			raw_saturation = row.get("sg")
		saturation = _to_float(raw_saturation, "saturation", i)
		for table_name in ordered_tables:
			raw_val = row.get(table_name)
			if raw_val is None or str(raw_val).strip() == "":
				continue
			value = _to_float(raw_val, table_name, i)
			tables.setdefault(table_name, []).append((saturation, value))

	for pairs in tables.values():
		pairs.sort(key=lambda item: item[0])
	return tables


def _to_float(raw_value: object, col_name: str, row_number: int) -> float:
	if raw_value is None:
		raise ValueError(f"Nilai kosong pada kolom {col_name}, baris {row_number}.")
	text = str(raw_value).strip()
	if not text:
		raise ValueError(f"Nilai kosong pada kolom {col_name}, baris {row_number}.")
	# Support decimal commas from spreadsheets.
	if "," in text and "." not in text:
		text = text.replace(",", ".")
	try:
		return float(text)
	except ValueError as exc:
		raise ValueError(f"Nilai tidak valid pada kolom {col_name}, baris {row_number}: {raw_value}") from exc


def _norm_col_name(name: object) -> str:
	text = "" if name is None else str(name).strip().lower()
	text = text.replace(" ", "_")
	text = text.replace("-", "_")
	return text


def load_example_rock_tables(project_config: ProjectConfig) -> ProjectConfig:
	project_config.rock_tables = {
		"kro": [(0.0, 1.0), (0.3, 0.75), (1.0, 0.0)],
		"krw": [(0.0, 0.0), (0.3, 0.05), (1.0, 1.0)],
		"krg": [(0.0, 0.0), (0.2, 0.04), (1.0, 1.0)],
		"pcow": [(0.0, 12.0), (0.3, 7.0), (1.0, 0.0)],
		"pcgw": [(0.0, 9.0), (0.2, 5.0), (1.0, 0.0)],
	}
	project_config.is_dirty = True
	return project_config


def clear_rock_tables(project_config: ProjectConfig) -> ProjectConfig:
	project_config.rock_tables.clear()
	project_config.is_dirty = True
	return project_config


def update_perturbation_config(
	project_config: ProjectConfig, pert: PerturbationConfig
) -> ProjectConfig:
	project_config.perturbation = pert
	project_config.constraints.perturbation_confirmed = pert.perturbed_cell_id > 0
	project_config.is_dirty = True
	return project_config


def update_method_config(project_config: ProjectConfig, method_config: MethodConfig) -> ProjectConfig:
	project_config.methods = method_config
	project_config.constraints.methods_confirmed = True
	project_config.is_dirty = True
	return project_config


def update_wells(project_config: ProjectConfig, wells: list[WellConfig]) -> ProjectConfig:
	project_config.wells = [w for w in wells if isinstance(w, WellConfig)]
	project_config.constraints.wells_confirmed = True
	project_config.is_dirty = True
	return project_config


def mark_project_clean(project_config: ProjectConfig) -> ProjectConfig:
	project_config.is_dirty = False
	return project_config


def save_project_json(project_config: ProjectConfig, file_path: str | Path) -> Path:
	target = write_project_json(project_config, file_path)
	mark_project_clean(project_config)
	return target


def load_project_json(file_path: str | Path) -> ProjectConfig:
	project = io_load_project_json(file_path)
	mark_project_clean(project)
	return project
